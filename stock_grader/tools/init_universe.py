#!/usr/bin/env python3
"""
init_universe.py -- generate config/universe.yaml from your existing workbook.

Reads 'Individual grading', pulls every ticker and its pinned archetype from
each block header, and writes a ready-to-use universe.yaml. Saves typing 90
tickers by hand and guarantees the block order matches the spreadsheet, which
is what the diff engine uses to compute cell addresses.

Usage:
    python tools/init_universe.py /path/to/Fundamental_Scoring_PTs.xlsx
    python tools/init_universe.py wb.xlsx --email you@yourfirm.com
    python tools/init_universe.py wb.xlsx --out config/universe.yaml --force
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook                                  # noqa: E402

BLOCK_HEIGHT = 13
SHEET = "Individual grading"

# Maps the workbook's archetype labels to the single letters the rubric uses.
LABEL_TO_LETTER = {
    "a: mature cash generator": "A",
    "b: durable compounder": "B",
    "c: high growth": "C",
    "d: cyclical / commodity": "D",
    "d: cyclical/commodity": "D",
    "e: early biotech": "E",
    "e: early stage": "E",          # the MBAI typo
    "f: financial institution": "F",
    "g: turnaround": "G",
}


def read_blocks(path: str) -> list:
    """Return [(ticker, archetype_letter, raw_label, block_index), ...]."""
    wb = load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET]

    out, empty_run = [], 0
    for idx in range(1, 251):
        hdr = 1 + BLOCK_HEIGHT * (idx - 1)
        ticker = ws.cell(row=hdr + 1, column=1).value
        label = ws.cell(row=hdr, column=10).value

        if not (isinstance(ticker, str) and ticker.strip()):
            empty_run += 1
            if empty_run >= 12 and out:
                break
            continue
        empty_run = 0

        raw = str(label).strip() if label else ""
        letter = LABEL_TO_LETTER.get(raw.lower())
        out.append((ticker.strip().upper(), letter, raw, idx))
    wb.close()
    return out


def build_yaml(blocks: list, email: str) -> str:
    L = [
        "# =============================================================================",
        f"# universe.yaml -- generated from the workbook on {datetime.now():%Y-%m-%d}",
        "# =============================================================================",
        "# Ticker order MATCHES the workbook's block order. Do not reorder them: the",
        "# diff engine derives cell addresses from position (block n -> row 1+13*(n-1)).",
        "",
        "settings:",
        "  # SEC EDGAR requires a real contact address or it rate-limits/blocks you.",
        f'  sec_user_agent: "Research {email}"',
        '  base_currency: "USD"',
        "",
        "  # Re-grade triggers",
        "  earnings_window_days: 5",
        "  fy2_move_trigger: 0.10",
        "  price_drawdown_trigger: -0.20",
        "  volume_spike_trigger: 3.0",
        "  max_snapshot_age_days: 100",
        "",
        "  # Tier 3 (LLM) re-extraction cadence",
        "  tier3_max_age_days: 400",
        "",
        "  # Peer discovery",
        "  max_peers: 8",
        "  min_peers_to_score: 3",
        "  peer_cache_days: 90",
        "",
        "  # LLM",
        '  llm_model: "claude-sonnet-4-6"',
        "  llm_enabled: true",
        "",
        "# -----------------------------------------------------------------------------",
        "# TICKERS -- archetypes pinned from the workbook's column J.",
        "# Add `peers: [A, B, C]` under any ticker to override automatic discovery.",
        "# -----------------------------------------------------------------------------",
        "tickers:",
        "",
    ]

    unknown = []
    for ticker, letter, raw, idx in blocks:
        L.append(f"  {ticker}:                      # block {idx}")
        if letter:
            L.append(f"    archetype: {letter}       # {raw}")
        else:
            unknown.append((ticker, raw))
            L.append(f"    # archetype: ?           # UNRECOGNISED LABEL: '{raw}'")
            L.append(f"    #   left unpinned -- the decision tree will classify it")
        L.append("")

    L += [
        "# -----------------------------------------------------------------------------",
        "# MANUAL OVERRIDES -- your judgment beats the machine and is never overwritten.",
        "# -----------------------------------------------------------------------------",
        "overrides: {}",
        "",
        "#  Example:",
        "#  overrides:",
        "#    AAOI:",
        "#      business_quality:",
        "#        score: 7",
        '#        note: "In-house MBE/MOCVD fab. Offset: Microsoft 28% of DC revenue."',
        '#        asof: "2026-01-15"',
        "",
    ]
    return "\n".join(L), unknown


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", help="path to Fundamental_Scoring_PTs.xlsx")
    ap.add_argument("--out", default=None, help="output path (default config/universe.yaml)")
    ap.add_argument("--email", default="yourname@yourfirm.com",
                    help="contact address for the SEC User-Agent header")
    ap.add_argument("--force", action="store_true", help="overwrite an existing file")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        raise SystemExit(f"Workbook not found: {args.workbook}")

    out = args.out or os.path.join(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__))), "config", "universe.yaml")

    blocks = read_blocks(args.workbook)
    if not blocks:
        raise SystemExit("No ticker blocks found -- is this the right workbook?")

    text, unknown = build_yaml(blocks, args.email)

    if os.path.exists(out) and not args.force:
        backup = out + f".backup_{datetime.now():%Y%m%d_%H%M%S}"
        os.rename(out, backup)
        print(f"  Existing config backed up to: {os.path.basename(backup)}")

    with open(out, "w", encoding="utf-8") as f:
        f.write(text)

    from collections import Counter
    counts = Counter(b[1] or "unpinned" for b in blocks)

    print(f"\n  Wrote {out}")
    print(f"  {len(blocks)} tickers, in workbook block order.\n")
    print("  Archetypes:")
    for k in sorted(counts):
        print(f"    {k:<9} {counts[k]}")
    if unknown:
        print(f"\n  {len(unknown)} unrecognised archetype label(s) -- left unpinned:")
        for tk, raw in unknown:
            print(f"    {tk:<7} '{raw}'")
    if args.email == "yourname@yourfirm.com":
        print("\n  NEXT: set a real contact address in settings.sec_user_agent")
        print("        (EDGAR blocks requests without one)")
    print()
    return 0


if __name__ == "__main__":
    sys.exit(main())
