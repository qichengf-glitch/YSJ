"""
excel.py -- write scores back into Fundamental_Scoring_PTs.xlsx.

Safety rules, in order of importance:
  1. Always write to a COPY. The original is never touched.
  2. Never overwrite a category that has a manual override in universe.yaml.
  3. Never touch column K (the =C*J weight products), the total formula, or the
     Model sheet -- those are the workbook's own arithmetic.
  4. Verify the ticker in column A matches before writing to a block. A silent
     off-by-one in the block index would write NVDA's scores onto MSFT.

Also offers `fix_workbook_formulas`, which repairs the two issues found in the
existing file: the hardcoded total formulas in the ABVX and AKAM blocks, and
hand-typed archetype weights in column J.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

from .state import BLOCK_HEIGHT, block_header_row
from .util import LOG

SHEET = "Individual grading"
TICKER_COL = 1      # A
CATEGORY_COL = 2    # B
SCORE_COL = 3       # C
REASON_COL = 4      # D
WEIGHT_COL = 10     # J
PRODUCT_COL = 11    # K


def discover_blocks(path: str, sheet: str = SHEET,
                    max_blocks: int = 250, empty_tolerance: int = 12) -> dict:
    """
    Map TICKER -> 1-based block index by reading column A of each block header.

    Note: `read_only=True` leaves `ws.max_row` as None for many writers, so the
    scan cannot rely on it. Instead it walks blocks and stops after a run of
    consecutive empty ones -- which also tolerates the blank template blocks
    that sit at the end of the sheet.
    """
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"sheet '{sheet}' not found; available: {wb.sheetnames}")
    ws = wb[sheet]

    out: dict = {}
    consecutive_empty = 0
    for idx in range(1, max_blocks + 1):
        hdr = block_header_row(idx)
        tk = ws.cell(row=hdr + 1, column=TICKER_COL).value
        if isinstance(tk, str) and tk.strip() and not tk.strip().lower().startswith("company"):
            out[tk.strip().upper()] = idx
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            if consecutive_empty >= empty_tolerance and out:
                break
    wb.close()
    LOG.info("discovered %d ticker blocks in %s", len(out), os.path.basename(path))
    return out


def write_scores(workbook_path: str, results: dict, block_index: dict,
                 overrides: Optional[dict] = None, output_path: Optional[str] = None,
                 write_reasons: bool = True, dry_run: bool = False) -> str:
    """Apply scores into a copy of the workbook. Returns the new file path."""
    overrides = overrides or {}
    output_path = output_path or workbook_path.replace(
        ".xlsx", f"_updated_{datetime.now().strftime('%Y%m%d')}.xlsx")

    if not dry_run:
        shutil.copy2(workbook_path, output_path)
    target = output_path if not dry_run else workbook_path

    wb = load_workbook(target, data_only=False)
    if SHEET not in wb.sheetnames:
        raise ValueError(f"sheet '{SHEET}' not found in {target}")
    ws = wb[SHEET]

    written = skipped = mismatched = 0

    for ticker, res in results.items():
        bi = block_index.get(ticker.upper())
        if bi is None:
            LOG.warning("[%s] no block in workbook -- skipped", ticker)
            continue
        hdr = block_header_row(bi)

        # Guard against writing into the wrong block.
        sheet_ticker = ws.cell(row=hdr + 1, column=TICKER_COL).value
        if not isinstance(sheet_ticker, str) or sheet_ticker.strip().upper() != ticker.upper():
            LOG.error("[%s] block %d holds '%s' -- REFUSING to write",
                      ticker, bi, sheet_ticker)
            mismatched += 1
            continue

        tk_overrides = {k.lower() for k in (overrides.get(ticker.upper(), {}) or {})}

        for ckey, cat in res.categories.items():
            if cat.score is None:
                continue
            if ckey in tk_overrides or cat.source == "manual":
                skipped += 1
                continue

            row = hdr + cat.sheet_row_offset
            if not dry_run:
                ws.cell(row=row, column=SCORE_COL).value = int(cat.score)
                if write_reasons:
                    bits = [s.audit().strip() for s in cat.subscores if s.status == "scored"]
                    caps = ["CAP: " + c for c in cat.caps_applied]
                    txt = " | ".join(bits + caps)
                    stamp = datetime.now().strftime("%Y-%m-%d")
                    ws.cell(row=row, column=REASON_COL).value = (
                        f"[auto {stamp}, conf={cat.confidence}] {txt}")[:2000]
            written += 1

    if not dry_run:
        wb.save(output_path)
    wb.close()

    LOG.info("scores written: %d applied, %d preserved as manual overrides, "
             "%d block mismatches%s", written, skipped, mismatched,
             " (DRY RUN -- nothing saved)" if dry_run else f" -> {output_path}")
    return output_path


def fix_workbook_formulas(workbook_path: str, output_path: Optional[str] = None,
                          weight_matrix_sheet: str = SHEET) -> str:
    """
    Repair the two defects in the existing workbook:

      1. ABVX (row 38) and AKAM (row 51) hardcode the BASE rubric weights in
         their total formula instead of using =SUM(K..)*10, so their archetype
         weights are ignored. ABVX scores 55 where its Early Biotech weights
         give 63.
      2. Column J weights are hand-typed. Replaced with an INDEX/MATCH against
         the archetype matrix at M36:T47 so they can never drift again.
    """
    output_path = output_path or workbook_path.replace(".xlsx", "_fixed.xlsx")
    shutil.copy2(workbook_path, output_path)

    wb = load_workbook(output_path, data_only=False)
    ws = wb[weight_matrix_sheet]

    # Archetype labels must match the matrix headers at N36:T36 exactly, or
    # MATCH fails and IFERROR silently zeroes every weight in the block.
    # MBAI is labelled "E: Early Stage" -- the same typo the Model sheet works
    # around with an IF remap. Repair it at the source instead.
    label_fixes = {
        "e: early stage": "E: Early Biotech",
        "e: early-stage": "E: Early Biotech",
        "a: mature cash generator ": "A: Mature Cash Generator",
    }
    valid_labels = {str(ws.cell(row=36, column=c).value).strip()
                    for c in range(14, 21)
                    if ws.cell(row=36, column=c).value}

    fixed_totals = fixed_weights = fixed_labels = 0
    idx = 1
    while True:
        hdr = block_header_row(idx)
        if hdr + BLOCK_HEIGHT > ws.max_row + BLOCK_HEIGHT:
            break
        tk = ws.cell(row=hdr + 1, column=TICKER_COL).value
        if not (isinstance(tk, str) and tk.strip()):
            idx += 1
            if idx > 200:
                break
            continue

        # ---- 0. archetype label must match a matrix header
        lbl_cell = ws.cell(row=hdr, column=WEIGHT_COL)          # J on the header row
        lbl = str(lbl_cell.value or "").strip()
        if lbl and lbl not in valid_labels:
            repl = label_fixes.get(lbl.lower())
            if repl:
                LOG.info("[%s] archetype label '%s' -> '%s'", tk.strip(), lbl, repl)
                lbl_cell.value = repl
                fixed_labels += 1
            else:
                LOG.error("[%s] archetype label '%s' matches no matrix header %s "
                          "-- weights would zero out; FIX THIS MANUALLY",
                          tk.strip(), lbl, sorted(valid_labels))

        # ---- 1. total formula
        total_row = hdr + 11
        cur = ws.cell(row=total_row, column=SCORE_COL).value
        want = f"=SUM(K{hdr+1}:K{hdr+10})*10"
        if isinstance(cur, str) and not cur.replace(" ", "").upper().startswith("=SUM(K"):
            LOG.info("[%s] total formula was hardcoded -- replacing with %s",
                     tk.strip(), want)
            ws.cell(row=total_row, column=SCORE_COL).value = want
            fixed_totals += 1

        # ---- 2. weights via INDEX/MATCH on the archetype matrix
        for off in range(1, 11):
            r = hdr + off
            ws.cell(row=r, column=WEIGHT_COL).value = (
                f"=IFERROR(INDEX($N$37:$T$46,{off},"
                f"MATCH(TRIM($J${hdr}),$N$36:$T$36,0)),0)")
            ws.cell(row=r, column=PRODUCT_COL).value = f"=C{r}*J{r}"
            fixed_weights += 1

        idx += 1
        if idx > 200:
            break

    wb.save(output_path)
    wb.close()
    LOG.info("workbook repaired: %d total formulas, %d weight cells, "
             "%d archetype labels -> %s",
             fixed_totals, fixed_weights, fixed_labels, output_path)
    return output_path
